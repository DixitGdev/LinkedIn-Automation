from openpyxl import load_workbook


def read_data():
    wb = load_workbook("./data/ACCOUNT TRACKER  - Raulene.xlsx")
    sheet = wb.active

    data_ls = []
    for row in sheet.iter_rows(min_row=7, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
        tmp = [each.value for each in row]
        if tmp[0]:
            data_ls.append(tmp)
    wb.close()
    return data_ls
